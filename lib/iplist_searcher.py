# -*- coding: utf-8 -*-
"""
Get target hostname's ipaddress list.
"""

from __future__ import (division, print_function, absolute_import,
                        unicode_literals)

from datetime import datetime
import os
import pickle

from openpyxl import load_workbook
import yaml


try:
    CastObj = unicode
except NameError:
    CastObj = str


# If YAML loading warning is printed, you can set False option.
# yaml.warnings({'YAMLLoadWarning': False})

class HostInfo(object):
    """
    Target host name and IPaddress.
    """

    def __init__(self, name, ipaddress):
        self._name = name
        self._ipaddress = ipaddress

    @property
    def name(self):
        """
        hostname
        """
        return self._name

    @property
    def ipaddress(self):
        """
        IPaddress
        """
        return self._ipaddress

    @name.setter
    def name(self, val):
        self._name = val

    @ipaddress.setter
    def ipaddress(self, val):
        self._ipaddress = val


class IPlistSearcher(object):
    """
    Generate all hostinfo from all of IPlist.
    """

    def __init__(self):
        _topdir = os.path.join(os.path.dirname(
            os.path.abspath(__file__)), '..')

        self._cachefile = os.path.join(_topdir, 'cachefile')
        self._iplistdir = os.path.join(_topdir, 'listfiles')

        with open('./conf.yaml') as yamlfile:
            conf = yaml.load(yamlfile)

        self.manage_columns = conf['manage_columns']

    @staticmethod
    def _get_sheet_hostinfo(worksheet, manlan_column):
        """ Collect hostinfo from every line. """

        def _check_multibyte_string(string):
            try:
                string.encode('ascii', 'strict')
                return True
            except UnicodeEncodeError:
                return False

        hostinfos = []
        for idx, _ in enumerate(worksheet['A']):
            line = str(idx + 1)
            cell = 'A' + line
            cellval = worksheet[cell].value

            if isinstance(cellval) != CastObj:
                cellval = CastObj(cellval)

            if cellval and _check_multibyte_string(cellval):
                # Mach hostname exists in this sheet.
                ip_column = manlan_column + line
                ipaddress = worksheet[ip_column].value
                if not ipaddress:
                    continue

                hostinfo = HostInfo(cellval, ipaddress)
                hostinfos.append(hostinfo)

        # This sheet's hostinfos.
        return hostinfos

    def _get_manlan_column(self, worksheet):
        for row in worksheet.rows:
            for idx, cell in enumerate(row):
                if cell.value and cell.value in self.manage_columns:
                    return chr(65 + idx)
                else:
                    continue

        return False

    def _make_host_list(self, iplist_path):
        """ Collect all hostinfo from every sheets. """
        iplist_hostinfos = []

        workbook = load_workbook(iplist_path)
        for sheetname in workbook.sheetnames:
            worksheet = workbook[sheetname]
            manlan_column = self._get_manlan_column(worksheet)
            if not manlan_column:
                continue

            hostinfos = self._get_sheet_hostinfo(worksheet, manlan_column)
            if hostinfos:
                iplist_hostinfos.append(hostinfos)

        return iplist_hostinfos

    @staticmethod
    def _initial_cache():
        return {'hostfile': {},
                'filetimes': {},
                'addresses': {}}

    def read_cache(self):
        """
        Read IPlist cache.
        """

        if not os.path.isfile(self._cachefile):
            return self._initial_cache()

        with open(self._cachefile, 'rb') as cachefile:
            cachedata = pickle.load(cachefile)

        return cachedata

    def write_cache(self, cachedata):
        """
        Write IPlist cache.
        """

        with open(self._cachefile, 'wb') as cachefile:
            pickle.dump(cachedata, cachefile)

    def _compare_cache(self, current_caches):
        """
        # Data structure
        current_caches = {
            'hostfile': {},
            'filetimes': {
                'iplist1.xlsx': '2018xxxx',
                'iplist2.xlsx': '2018xxxx'
            },
            'addresses': {
                'iplist1.xlsx': [[HostInfo(name, ipaddress),
                                  HostInfo(name, ipaddress)]],
                'iplist2.xlsx': [[HostInfo(name, ipaddress),
                                  HostInfo(name, ipaddress)]]
            }
        }

        # Create cache data
        hostnames = {}
        for filename in current_caches['filetimes'].keys():
            for hostname in current_caches['addresses']['filename'].keys():
                hostnames[hostname] = filename

        current_caches['hostfile'] = hostnames

        # Usage
        filename = current_caches['hostfile']['targethost1']
        address = current_caches['addresses']['filename']['targethost']
        """

        files = os.listdir(self._iplistdir)
        files = [iplist for iplist in files
                 if iplist.endswith('.xlsx') or iplist.endswith('.xls')]

        current_filetimes = {}
        for filename in files:
            unixtime = os.stat(self._iplistdir + '/' + filename).st_mtime
            timestamp = datetime.fromtimestamp(
                unixtime).strftime('%Y%m%d%H%M%S')
            current_filetimes[filename] = timestamp

        replaced = False
        new_cache_part = self._initial_cache()
        for filename in current_filetimes.keys():
            if filename not in current_caches['filetimes']\
               or (current_caches['filetimes'][filename] !=
                   current_filetimes[filename]):

                replaced = True
                new_cache_part['filetimes'][filename] = \
                    current_filetimes[filename]
            else:
                new_cache_part['filetimes'][filename] = \
                    current_caches['filetimes'][filename]
                if filename in current_caches['addresses']:
                    new_cache_part['addresses'][filename] = \
                        current_caches['addresses'][filename]

        if replaced and current_caches['hostfile']:
            current_caches['hostfile'] = {}
        else:
            new_cache_part['hostfile'] = current_caches['hostfile']

        return new_cache_part, replaced

    def create_cache(self, current_caches):
        """
        Create cache from IPlists.
        """

        new_cache_part, replaced = self._compare_cache(current_caches)

        if not replaced:
            # new_cache_part == current_caches
            return new_cache_part

        print('Rebuilding cache data...')
        replaced_files = list(set(new_cache_part['filetimes'].keys())
                              - set(new_cache_part['addresses'].keys()))

        for iplist in list(replaced_files):
            iplist_path = os.path.join(self._iplistdir, iplist)
            if not os.path.isfile(iplist_path):
                new_cache_part['filetimes'].pop(iplist)
                continue

            # create hostinfos by all rows.
            iplist_hostinfos = self._make_host_list(iplist_path)
            if iplist_hostinfos:
                new_cache_part['addresses'][iplist] = iplist_hostinfos

        hostnames = {}
        for filename in new_cache_part['addresses'].keys():
            for sheet in new_cache_part['addresses'][filename]:
                for hostinfo in sheet:
                    hostnames[hostinfo.name] = filename

        new_cache_part['hostfile'] = hostnames

        return new_cache_part


def get_all_hostinfos():
    """ Get all host's ipaddress from all iplist files. """

    searcher = IPlistSearcher()

    current_caches = searcher.read_cache()
    new_caches = searcher.create_cache(current_caches)
    searcher.write_cache(new_caches)

    all_iplist_hostinfos = new_caches['addresses']
    return all_iplist_hostinfos


def get_target_hostinfo(server_name):
    """
    Get target host info.
    """

    return HostInfo(server_name, server_name)
